#!/usr/bin/env python3
"""Build the 18-sequence main rendering table (Method x Sequence x {ATE, PSNR, SSIM,
LPIPS, Depth L1}) combining:
  * Our columns: combined(mask-ON) and mask-free — ATE from tracking_raw, rendering
    from posthoc_fullframe (3090 offline re-render of final/ PLY), 3-seed mean±std.
  * Competitor columns: from resources/02-baselines/04-baselines_result.xlsx
    (01_ATE_RMSE, 06_PSNR, 07_SSIM, 08_LPIPS, 09_Depth_L1), mean only for readability.
Emits results/evidence/18seq_rendering_main_table.md (paper table).
"""
import csv, glob, os, statistics, json

# (run_dir, [completed timestamps]) for seeds that were run more than once; reported
# in the table footer so a duplicate can never pass silently again.
DUPLICATE_RUNS = []

ROOT = "results/runs"
SEQORDER = ["f1_desk","f2_xyz","f3_office","f2_person","f3_st_hf","f3_st_rpy",
            "f3_st_xyz","f3_wk_hf","f3_wk_rpy","f3_wk_xyz","balloon","balloon2",
            "crowd","crowd2","mv_no_box","mv_no_box2","pt1","pt2"]
PTYPE = {
    "f1_desk":"TUM 静态","f2_xyz":"TUM 静态","f3_office":"TUM 静态","f2_person":"TUM 动态",
    "f3_st_hf":"TUM sitting","f3_st_rpy":"TUM sitting","f3_st_xyz":"TUM sitting",
    "f3_wk_hf":"TUM walking","f3_wk_rpy":"TUM walking","f3_wk_xyz":"TUM walking",
    "balloon":"BONN 混合","balloon2":"BONN 混合","crowd":"BONN 多人","crowd2":"BONN 多人",
    "mv_no_box":"BONN 纯物","mv_no_box2":"BONN 纯物","pt1":"BONN 纯人","pt2":"BONN 纯人"}

P6MASON_COMBINED = {"crowd","crowd2","f3_wk_rpy","f3_wk_xyz","pt1"}

# ---- FULLKERN rerun: the 11 sequences whose ORIGINAL main-table runs were silent K1R1L0 --
# INCIDENT (exp23/exp24). These 11 sequences had no `flow_raft/` precomputed, so
# ReliabilitySignal silently no-op'd: the runs are labelled combined/mask-free but were
# actually missing the L component (K1R1L0, not K1R1L1). A runtime gate now aborts instead
# of no-op'ing (utils/reliability_signal.py::assert_reliability_flow_available, 7b89ff81),
# flow was built for all 11, and BOTH arms were rerun 3-seed on jiangwenheng.
#   combined(mask-ON) -> P6-FULLKERN            (was P6-MASON x3 + P6-MASON-8SEQ x8)
#   mask-free         -> P6-FULLKERN-MASKFREE   (was P6-18SEQ x11)
# See results/evidence/reliability_signal_silent_noop_incident.md.
#
# PRECEDENCE IS EXPLICIT, NOT DICT-ORDER: for these sequences the OLD roots are refused
# outright and the FULLKERN roots are the only admissible source. Coverage is then
# asserted (assert_fullkern_coverage) so a half-landed rerun can never silently fall back
# to the contaminated runs, nor silently drop the row from the table.
FULLKERN_SEQS = {"crowd","crowd2","f3_wk_rpy","f1_desk","f2_person","f3_office",
                 "f3_st_hf","f3_st_rpy","f3_st_xyz","f3_wk_hf","f2_xyz"}
FULLKERN_ROOTS = {
    "results/runs/P6/P6-FULLKERN/*": "maskon",
    "results/runs/P6/P6-FULLKERN-MASKFREE/*": "maskfree",
}
FULLKERN_SEEDS = ("0","1","2")
# (seq, kind) -> source label, for the table's provenance footer.
SRC_OF = {}

# ---- WP-B naive flow-threshold baseline (p90) ----------------------------------
# Middle-ground baseline demanded by review R2: same frozen offline RAFT flow budget as
# MRCS, consumed by a naive per-pixel threshold mask instead of our mechanism. Verdict B1
# (results/evidence/wpb_flowmask_verdict.md, 36-run confirm campaign).
# Scope is the WP-B held-out set ONLY — the arm was never run on the other 14 sequences,
# so they stay N/A here (prereg forbids back-filling from dev sequences or extrapolating).
# No rendering pass exists for this arm -> PSNR/SSIM/LPIPS/Depth-L1 stay "—".
WPB_FLOWMASK_ROOT = "results/runs/WPB/WPB-CONFIRM"

# ---- vanilla (MonoGS) 3-seed dispersion ----------------------------------------
# The competitor xlsx carries a single MonoGS mean per sequence, which is what let an
# earlier draft quote improvement ratios with no baseline dispersion at all. But this
# project HAS a same-environment 3-seed MonoGS rerun set, and it is the denominator the
# ratio audit fixed as the only legal one (results/evidence/headline_ratio_recompute.md).
# Read it here so the table itself shows the spread: 8 of the 18 sequences have vanilla
# CV > 20% -- mv_no_box 62% and mv_no_box2 101% among them -- and a reader must be able
# to see that WITHOUT opening another file, because a ratio built on those denominators
# is not well defined.
VANILLA_3SEED_CSV = "resources/02-baselines/baselines_result/MonoGS/tracking_raw.csv"
# CV above which a ratio may not be quoted as a single number (audit §0 / roadmap §五-A).
VANILLA_CV_QUOTABLE = 0.20
# The baseline CSV uses the BONN upstream names for the two person sequences while this
# table (and the whole paper) uses pt1/pt2. Mapping is asserted, not assumed: the means it
# produces must reproduce headline_ratio_recompute.md to 0.01 cm, checked below.
VANILLA_SEQ_ALIAS = {"pt1": "person_t", "pt2": "person_t2"}
# (seq -> expected mean, sd) from results/evidence/headline_ratio_recompute.md §1, so a
# renamed or re-landed baseline CSV can never silently change the ratio denominator.
VANILLA_EXPECTED = {
    "pt1": (44.83, 9.06), "pt2": (43.85, 8.55), "balloon": (39.32, 1.01),
    "mv_no_box": (15.33, 9.47), "mv_no_box2": (16.84, 16.96),
    "balloon2": (22.05, 1.55), "f3_wk_xyz": (28.14, 0.86),
    "f3_wk_rpy": (62.89, 6.81), "f3_wk_hf": (44.45, 11.06),
    "crowd": (86.47, 18.34), "crowd2": (147.46, 39.03),
}

WPB_FLOWMASK_SEQS = ["pt1","pt2","mv_no_box2","balloon2"]

def discover():
    out = {}
    roots = {
        "results/runs/P6/P6-18SEQ/*": "maskfree",
        "results/runs/P6/P6-MASKOFF-3SEED/*": "maskfree",
        "results/runs/P6/P6-MASKOFF/*": "maskfree",
        "results/runs/P2/P2-T_3090/*_prune_seed*": "maskon",
        "results/runs/P6/P6-MASON/*": "maskon",
        # missing8 combined(mask-ON) backfill: f1_desk/f2_xyz/f3_office/f2_person/
        # f3_st_{hf,rpy,xyz}/f3_wk_hf x3 seed. Not in P6MASON_COMBINED -> no src-conflict.
        "results/runs/P6/P6-MASON-8SEQ/*_combined_seed*": "maskon",
    }
    # FULLKERN reruns are listed last but are NOT order-dependent: the loop below refuses
    # the old roots for FULLKERN_SEQS outright, so there is exactly one admissible source
    # per (seq, kind) and no last-writer-wins race.
    roots.update(FULLKERN_ROOTS)
    for pat, kind in roots.items():
        is_fullkern = pat in FULLKERN_ROOTS
        for d in glob.glob(pat):
            if not os.path.isdir(d) or "consolelog" in os.path.basename(d): continue
            nm = os.path.basename(d)
            mm = nm.rsplit("_seed", 1)
            if len(mm) != 2: continue
            runname, seed = mm[0], mm[1]
            seq = src = None
            if "_maskoff" in runname: seq=runname.replace("_maskoff",""); src="P6-MASKOFF"
            elif "_prune" in runname: seq=runname.replace("_prune",""); src="P2-T-3090"
            elif "_combined" in runname: seq=runname.replace("_combined",""); src="P6-MASON"
            else: continue
            # --- silent-no-op incident quarantine (see FULLKERN_SEQS above) ---
            if is_fullkern:
                # the FULLKERN dirs only ever hold the 11 reran sequences; refuse anything
                # else so a stray run there can't enter the table unnoticed.
                if seq not in FULLKERN_SEQS: continue
                src = "P6-FULLKERN" if kind=="maskon" else "P6-FULLKERN-MASKFREE"
            elif seq in FULLKERN_SEQS:
                continue  # contaminated K1R1L0 original -> never admissible
            if kind=="maskon" and seq in P6MASON_COMBINED and src=="P2-T-3090":
                continue  # P6-MASON is canonical mask-ON for pt1/crowd/wk_*
            latest=None; completed=[]
            for ts in glob.glob(d+"/datasets_*/*/seed_*/*/"):
                ts=ts.rstrip("/")
                if not os.path.isfile(ts+"/config.yml"): continue
                if not os.path.exists(ts+"/posthoc_fullframe/fullframe_summary.json"): continue
                completed.append(ts)
                if latest is None or ts>latest: latest=ts
            if latest is None: continue
            # never silent: a seed that was run more than once is a provenance decision,
            # not a detail. We keep the LATEST (rendering + ATE then agree) and say so.
            # NOTE: `completed` only counts runs that were ALSO re-rendered (posthoc
            # summary present). The f3_wk_rpy case had a second FULL run that was never
            # re-rendered, so it is invisible here -- which is exactly how the old
            # roll-up averaging slipped through. The roll-up row count catches it.
            rollup=os.path.join(d,"tables","tracking_raw.csv")
            n_rollup=0; ids=[]
            if os.path.isfile(rollup):
                try:
                    with open(rollup) as f:
                        rows=[r for r in csv.DictReader(f) if (r.get("ate_rmse_cm") or "").strip()]
                    n_rollup=len(rows); ids=[(r.get("run_id") or "?").strip() for r in rows]
                except Exception: pass
            if len(completed)>1 or n_rollup>1:
                DUPLICATE_RUNS.append((nm, ids or [os.path.basename(t) for t in sorted(completed)]))
            out.setdefault(seq,{}).setdefault(kind,{})[seed]=latest
            SRC_OF[(seq,kind)]=src
    assert_fullkern_coverage(out)
    return out


def assert_fullkern_coverage(disc):
    """Refuse to emit a table unless the FULLKERN rerun fully replaced the 11 tainted seqs.

    Two failure modes this closes, both of which would otherwise be SILENT:
      1. rerun half-landed -> the row quietly drops out of the 18-seq table (a shorter
         table reads as 'that sequence was never run', not 'the rerun is incomplete');
      2. a FULLKERN run finished tracking but was never re-rendered -> `discover()` skips
         it for lack of posthoc_fullframe, same silent drop.
    Both look identical to success in the emitted markdown, which is exactly the class of
    failure the silent-no-op incident was. So: hard error, naming every missing cell.
    """
    missing=[]
    for seq in sorted(FULLKERN_SEQS):
        for kind in ("maskfree","maskon"):
            have=disc.get(seq,{}).get(kind,{})
            src=SRC_OF.get((seq,kind))
            for sd in FULLKERN_SEEDS:
                if sd not in have:
                    missing.append(f"{seq}/{kind}/seed{sd}")
            if have and src not in ("P6-FULLKERN","P6-FULLKERN-MASKFREE"):
                missing.append(f"{seq}/{kind}: WRONG SOURCE {src}")
    if missing:
        raise SystemExit(
            "FULLKERN coverage incomplete -- refusing to write the main table.\n"
            f"  {len(missing)} missing/invalid cell(s):\n    " + "\n    ".join(missing) +
            "\n  Each FULLKERN run needs BOTH tracking_raw.csv AND "
            "posthoc_fullframe/fullframe_summary.json (offline re-render).\n"
            "  See results/evidence/reliability_signal_silent_noop_incident.md.")

def read_ate(tsd):
    """ATE of the ONE run `tsd` -- never an average over the roll-up.

    PROVENANCE FIX (2026-08-15, exp22). The roll-up `<RUNROOT>/tables/tracking_raw.csv`
    is APPEND-ONLY and can hold more than one completed run of the same seed. This
    function used to average every row in it, so a seed that was run twice contributed
    the mean of both runs -- while the rendering columns above read only the LATEST
    timestamp. The two halves of such a row then came from different runs.
    Observed once in 258 runs: `f3_wk_rpy_maskoff_seed0` ran twice (873/873 frames both,
    ATE 15.72 vs 21.42 cm), which is also this project's clearest same-seed
    nondeterminism evidence -- see `results/evidence/main_table_provenance_audit.md`.
    Authoritative order: the per-run CSV inside the timestamp dir, else the roll-up row
    whose `run_id` matches that timestamp.
    """
    stamp=os.path.basename(tsd.rstrip(os.sep))
    rows=[]
    per_run=os.path.join(tsd,"tracking_raw.csv")
    try:
        if os.path.isfile(per_run):
            with open(per_run) as f: rows=list(csv.DictReader(f))
        else:
            parts=tsd.split(os.sep); runroot=os.sep.join(parts[:-4])
            csvp=os.path.join(runroot,"tables","tracking_raw.csv")
            if not os.path.isfile(csvp): return None
            with open(csvp) as f: allrows=list(csv.DictReader(f))
            rows=[r for r in allrows if (r.get("run_id") or "").strip()==stamp] or allrows[-1:]
        vals=[float(r["ate_rmse_cm"]) for r in rows if r.get("ate_rmse_cm","").strip()]
        if not vals: return None
    except Exception: return None
    return (round(vals[-1],2), None)

def read_summary(tsd):
    p=os.path.join(tsd,"posthoc_fullframe","fullframe_summary.json")
    if not os.path.isfile(p): return None
    try:
        with open(p) as f: return json.load(f).get("fullframe",{})
    except Exception: return None

def agg(tsds):
    keys=["psnr","ssim","lpips","depth_l1_cm"]
    stats={k:[] for k in keys}
    ates=[]
    for seed,tsd in tsds.items():
        ff=read_summary(tsd)
        if ff:
            for k in keys: stats[k].append(ff.get(k))
        ates.append(read_ate(tsd))
    def mean_std(vals, nd=2):
        vals=[v for v in vals if v is not None]
        if not vals: return "—"
        m=statistics.mean(vals)
        if len(vals)>1:
            sd=statistics.stdev(vals)
            return f"{m:.{nd}f}±{sd:.{nd}f}"
        return f"{m:.{nd}f}"
    af=[a for a in ates if a]
    ate=f"—"
    if af:
        am=statistics.mean(a[0] for a in af)
        asd=statistics.stdev(a[0] for a in af) if len(af)>1 else None
        ate=f"{am:.2f}"+ (f"±{asd:.2f}" if asd else "")
    return ate, mean_std(stats["psnr"]), mean_std(stats["ssim"],3), mean_std(stats["lpips"],3), mean_std(stats["depth_l1_cm"])

def read_flowmask():
    """3-seed ATE (mean±std, ddof=1) per held-out seq for the naive flow-threshold(p90) arm.

    Reads the same tracking_raw.csv chain as our own rows, so the numbers carry the same
    provenance. Dispersion is std here for consistency with the rest of this table; the
    WP-B verdict doc quotes the SAME runs as mean±half-range (a different spread measure,
    not a different dataset).
    """
    out={}
    for seq in WPB_FLOWMASK_SEQS:
        vals=[]
        for d in sorted(glob.glob(os.path.join(WPB_FLOWMASK_ROOT, f"flowmask_{seq}_*"))):
            if not os.path.isdir(d): continue
            csvp=os.path.join(d,"tables","tracking_raw.csv")
            if not os.path.isfile(csvp): continue
            try:
                with open(csvp) as f:
                    for r in csv.DictReader(f):
                        if r.get("status","").strip()!="OK": continue
                        v=r.get("ate_rmse_cm","").strip()
                        if v: vals.append(float(v))
            except Exception: continue
        if not vals: continue
        m=statistics.mean(vals)
        out[seq]=f"{m:.2f}"+(f"±{statistics.stdev(vals):.2f}" if len(vals)>1 else "")
    return out

# ---- load competitor xlsx ----
import openpyxl
wb=openpyxl.load_workbook("resources/02-baselines/04-baselines_result.xlsx",data_only=True)
def comp_sheet(sheet, colname):
    ws=wb[sheet]
    # method names are in col B
    out={}
    for r in range(2, ws.max_row+1):
        m=ws.cell(r,2).value
        if not m: continue
        seqvals={}
        for i,s in enumerate(SEQORDER):
            raw=ws.cell(r,4+i).value
            if raw is None or str(raw).strip() in ("","-","F","N/A"): seqvals[s]=None; continue
            try: seqvals[s]=float(str(raw).split()[0])
            except: seqvals[s]=None
        out[m]=seqvals
    return out
comp_ate=comp_sheet("01_ATE_RMSE","ate")
comp_psnr=comp_sheet("06_PSNR","psnr")
comp_ssim=comp_sheet("07_SSIM","ssim")
comp_lpips=comp_sheet("08_LPIPS","lpips")
comp_depth=comp_sheet("09_Depth_L1","depth")


def read_vanilla_3seed():
    """Per-sequence vanilla ATE as (mean, sd, n, cv) from the 3-seed MonoGS rerun set.

    Only status=OK rows count. A sequence with fewer than 2 usable seeds gets sd=None
    rather than a silent 0.0, so the table can never render a fabricated "±0.00".
    """
    by_seq = {}
    if not os.path.exists(VANILLA_3SEED_CSV):
        raise SystemExit(f"vanilla 3-seed denominator missing: {VANILLA_3SEED_CSV}")
    with open(VANILLA_3SEED_CSV, encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            if (row.get("status") or "").strip().upper() != "OK":
                continue
            try:
                by_seq.setdefault(row["sequence"], []).append(float(row["ate_rmse_cm"]))
            except (KeyError, TypeError, ValueError):
                continue
    out = {}
    for seq, vals in by_seq.items():
        mean = statistics.mean(vals)
        sd = statistics.stdev(vals) if len(vals) > 1 else None
        cv = (sd / mean) if (sd is not None and mean) else None
        out[seq] = (mean, sd, len(vals), cv)
    # resolve the pt1/pt2 aliases onto the paper's names
    for paper_name, csv_name in VANILLA_SEQ_ALIAS.items():
        if csv_name in out:
            out[paper_name] = out[csv_name]
    # the denominator is load-bearing for every ratio in the paper: verify it against the
    # audit's published values instead of trusting whatever CSV happens to be on disk.
    drift = []
    for seq, (exp_m, exp_sd) in VANILLA_EXPECTED.items():
        rec = out.get(seq)
        if rec is None:
            drift.append(f"{seq}: absent")
            continue
        got_m, got_sd = rec[0], rec[1]
        if abs(got_m - exp_m) > 0.011 or got_sd is None or abs(got_sd - exp_sd) > 0.011:
            drift.append(f"{seq}: csv=({got_m:.2f},{got_sd}) audit=({exp_m},{exp_sd})")
    if drift:
        raise SystemExit("vanilla denominator disagrees with headline_ratio_recompute.md, "
                         "refusing to emit table:\n  " + "\n  ".join(drift))
    return out


def fmt_vanilla(rec):
    """`mean±sd (CV%)`, flagging the ones whose ratio may not be quoted alone."""
    if rec is None:
        return "—"
    mean, sd, n, cv = rec
    m = f"{mean:.2f}"
    if sd is None or cv is None:
        return f"{m} (n={n})"
    flag = " ⚠" if cv > VANILLA_CV_QUOTABLE else ""
    return f"{m}±{sd:.2f} (CV {cv * 100:.0f}%){flag}"

COMP=["MonoGS","SplaTAM","Co-SLAM","DG-SLAM","WildGS-SLAM","RGD-SLAM"]
# headline competitor set for table (keep table tractable): MonoGS + RGD as SOTA two
HEADLINE=["MonoGS","RGD-SLAM"]

def main():
    # importable as a lib (write_18seq_main_xlsx.py); guarded so import doesn't rewrite md
    disc=discover()
    flowmask=read_flowmask()
    vanilla=read_vanilla_3seed()
    missing=[s for s in SEQORDER if s not in vanilla]
    if missing:
        raise SystemExit("vanilla 3-seed denominator incomplete, refusing to emit table: "
                         + ", ".join(missing))
    lines=[]
    lines.append("# 18 序列渲染主表（paper main table, 2026-08-11）")
    lines.append("")
    lines.append("> 来源标注: **我们** = 3090 离线重渲 `final/` PLY + `trj_full_final`（`posthoc_fullframe`，3-seed mean±std, ddof=1）; "
                 "**竞品** = 用户自测 `04-baselines_result.xlsx`（mean）。渲染 = 全帧 PSNR/SSIM/LPIPS/Depth-L1。"
                 "我们的渲染协议（`final/` PLY）与竞品同口径比较（均全分辨率全帧）。FPS 是 2060 数，不与竞品并列。")
    lines.append("")
    lines.append("> **⚠ MonoGS 行的 ATE 是 3-seed mean±sd (CV)，不是 xlsx 的单一均值**"
                 f"（2026-08-25 exp46 改）。源 = `{VANILLA_3SEED_CSV}`（同环境同口径重跑，18 序列 × 3 seed）。"
                 "这是 `headline_ratio_recompute.md` 定为**唯一合法**的倍数分母。"
                 f"标 ⚠ 的格子 vanilla CV > {VANILLA_CV_QUOTABLE*100:.0f}%"
                 " ⇒ **该序列的 improvement ratio 依 basin 而变，不得写成单一倍数**，"
                 "只能写 mean±sd 或给区间。渲染四列仍取自 xlsx（那里只有 mean），故 MonoGS 行"
                 "**ATE 有离散度、渲染没有** —— 这是取数来源不同，不是漏报。")
    lines.append("")
    lines.append("## 主表（Method × Sequence）")
    lines.append("")
    lines.append("| Seq | 类型 | 方法 | ATE(cm)↓ | PSNR↑ | SSIM↑ | LPIPS↓ | Depth-L1(cm)↓ |")
    lines.append("|---|---|---|---|---|---|---|---|")
    # For each seq, emit our mask-free + our mask-ON rows + competitor rows
    for seq in SEQORDER:
        ptype=PTYPE[seq]
        # our rows
        d=disc.get(seq,{})
        for kind,label in (("maskfree","Ours-mask-free"),("maskon","Ours-combined(mask-ON)")):
            if kind not in d or not d[kind]: continue
            ate,psnr,ssim,lpips,dep=agg(d[kind])
            lines.append(f"| {seq} | {ptype} | **{label}** | {ate} | {psnr} | {ssim} | {lpips} | {dep} |")
        # WP-B naive flow-threshold baseline: held-out 4 seqs only, ATE only (never rendered).
        if seq in flowmask:
            lines.append(f"| {seq} | {ptype} | Baseline-flow-mask(p90) | {flowmask[seq]} | — | — | — | — |")
        for m in HEADLINE:
            av=comp_ate.get(m,{}).get(seq)
            pv=comp_psnr.get(m,{}).get(seq)
            sv=comp_ssim.get(m,{}).get(seq)
            lv=comp_lpips.get(m,{}).get(seq)
            dv=comp_depth.get(m,{}).get(seq)
            def f(v, nd=2): return "—" if v is None else (f"{v:.{nd}f}" if v>=0 else "—")
            ate_c = "—" if av is None else f"{av:.2f}"
            # MonoGS is OUR backbone and the ratio denominator, so it is the one
            # competitor row that must carry dispersion rather than a bare mean.
            if m == "MonoGS":
                ate_c = fmt_vanilla(vanilla.get(seq))
                label = "MonoGS (vanilla, 3-seed)"
            else:
                label = m
            lines.append(f"| {seq} | {ptype} | {label} | {ate_c} | {f(pv)} | {f(sv,3)} | {f(lv,3)} | {f(dv)} |")
    lines.append("")
    # Coverage note
    lines.append("## 覆盖说明")
    lines.append("")
    lines.append("### 我们的方法")
    lines.append(f"- **mask-free**: 全 18 序列 × 3 seed（54 runs）渲染已齐。")
    lines.append(f"- **combined(mask-ON)**: 全 18 序列 × 3 seed 渲染已齐（原 10 序列 + missing8 补齐 f1_desk/f2_xyz/f3_office/f2_person/f3_st_{{hf,rpy,xyz}}/f3_wk_hf）。")
    lines.append(f"- 静态/低遮挡序列（f1_desk/f2_xyz/f3_office/f2_person/f3_st_*）mask-ON ≤ mask-free 或相当，掩码不形成缺口，作为竞争力支撑段。")
    lines.append("")
    lines.append(f"### ⚠ FULLKERN 重跑（{len(FULLKERN_SEQS)} 序列 × 2 臂 × 3 seed = {len(FULLKERN_SEQS)*6} run）")
    lines.append("")
    lines.append(f"- **为什么重跑**：这 {len(FULLKERN_SEQS)} 条序列的**原始主表 run 没有预计算 `flow_raft/`**，"
                 "`ReliabilitySignal` 因此被**静默跳过** —— 臂名写着 combined / mask-free，实跑是 "
                 "**K1R1L0**（缺 L 组件），不是 K1R1L1。这是错误的臂标，不是噪声。")
    lines.append(f"  序列：{', '.join(sorted(FULLKERN_SEQS))}。")
    lines.append("- **修复**：运行时硬闸（`utils/reliability_signal.py::assert_reliability_flow_available`，"
                 "commit `7b89ff81`）改为**缺 flow 直接 abort**，不再静默降级；补建全部 flow；两臂各 3 seed 重跑。")
    lines.append("- **本表取数**：这些序列**只**读 `P6-FULLKERN`（combined）/ `P6-FULLKERN-MASKFREE`（mask-free）。"
                 "旧的 `P6-18SEQ` / `P6-MASON` / `P6-MASON-8SEQ` 对应格**在代码层被拒绝**（非人工挑选），"
                 "覆盖不全时脚本**硬报错拒绝出表**，不会静默回落或静默少行。")
    lines.append("- **其余 7 序列**（balloon, balloon2, mv_no_box, mv_no_box2, pt1, pt2, f3_wk_xyz）"
                 "原本就有 flow，未受影响，保持原源。")
    lines.append("- **引用纪律**：任何跨事故前后的数字对比必须注明口径 —— 旧数是 K1R1L0，新数才是完整内核。"
                 "详见 `results/evidence/reliability_signal_silent_noop_incident.md`。")
    lines.append("")
    lines.append("### Baseline-flow-mask(p90)（WP-B 中间地带基线）")
    lines.append("")
    na = [s for s in SEQORDER if s not in flowmask]
    lines.append("- **是什么**：与 MRCS **共用同一套冻结离线 RAFT flow**，但用朴素逐像素阈值（`flow_quantile=0.9`，"
                 "阈值在 pilot 阶段用 dev 序列冻结）生成掩码，替代我们的机制。用于回答审稿 R2"
                 "「增益是否可归因于随便什么抗动态处理」。判决 **B1**，见 `wpb_flowmask_verdict.md`。")
    lines.append(f"- **有值范围 = WP-B held-out 4 序列**（{', '.join(WPB_FLOWMASK_SEQS)}），3 seed × 4 序列 = 12 run，"
                 "全部 `status=OK`；数据链与我方行同源（`results/runs/WPB/WPB-CONFIRM/*/tables/tracking_raw.csv`）。")
    lines.append(f"- **其余 {len(na)} 序列 = N/A**（{', '.join(na)}）：该臂**从未在这些序列上跑过**"
                 "（WP-B campaign 范围就是 held-out 4 序列），**非漏跑、非漏报**。"
                 "按预注册，**不得**为补格而外推或复用 dev 序列数据。")
    lines.append("- **渲染四列 = 「—」**：该臂只跑了 tracking，没有离线重渲，故无 PSNR/SSIM/LPIPS/Depth-L1。")
    lines.append("- **离散度口径**：本表统一 mean±std(ddof=1)；`wpb_flowmask_verdict.md` 对**同一批 run** "
                 "报 mean±half-range（口径不同，非数据不同）。")
    lines.append("- **公平性说明（引用本行必须同写；2026-08-15 更正）**：两臂共用的冻结 flow 是 "
                 "**backward `f_{t→t-1}`**（每帧只用该帧与前一帧）⇒ **信息上因果**；离线预计算只为"
                 "逐字节可复现 + 把 RAFT 移出 6GB 在线预算。因此本行支持「同等因果 flow 信息预算下，"
                 "朴素阈值 vs MRCS」的比较；仍成立的 caveat = **在线 FPS 不含 RAFT 推理开销**。"
                 "旧表述「双向/未来帧可见/非因果」已撤回，见 `flow_causality_correction.md`。")
    lines.append("- **不可与竞品列直接并读**：本行是我们自建的受控基线（同 campaign、同协议、同 flow 预算），"
                 "与表内 MonoGS/RGD-SLAM 的外部实测不是同一类比较。")
    lines.append("")
    lines.append("### 竞品")
    lines.append("- 表内 headlist = MonoGS（我们的基座）+ RGD-SLAM（动态 SOTA）。其余方法（SplaTAM/Co-SLAM/DG-SLAM/WildGS-SLAM/DynaGSLAM）在 `04-baselines_result.xlsx` 全 18 序列均有渲染，按需扩列。")
    lines.append("- `-` = 该 baseline 本身没有或不可比（论文设定里就没这指标），非漏跑。")
    lines.append("")
    lines.append("## 竞品渲染列（其余方法, 可在 xlsx 06-09 直接对号）")
    lines.append("")
    lines.append("PSNR: RGD 全 18 序列 19.0-25.1; WildGS 15.2-22.3; DG 13.1-24.2; MonoGS 13.8-25.0; SplaTAM 14.3-25.1; Co-SLAM 11.7-19.1。")
    lines.append("")
    lines.append("## 运行 provenance（自动检查，不可静默）")
    lines.append("")
    if DUPLICATE_RUNS:
        lines.append(f"⚠ **{len(DUPLICATE_RUNS)} 个 seed 目录含一个以上已完成 run**。本表对这些格"
                     "**只取最新那次**（ATE 与渲染因此同源于同一 run），旧次不并入均值：")
        for nm, stamps in sorted(DUPLICATE_RUNS):
            lines.append(f"  - `{nm}`：{len(stamps)} 次完成 —— {', '.join(stamps)}（取最新）")
        lines.append("  同 seed 重复运行之间的差异是**运行间非确定性**（MonoGS 前后端异步，"
                     "每帧 mapping 迭代数随 wall-clock 变化），详见 `main_table_provenance_audit.md`。")
    else:
        lines.append("✅ 每个 (序列, 臂, seed) 恰好一个已完成 run，ATE 与渲染同源。")
    lines.append("")
    out=os.path.join("results","evidence","18seq_rendering_main_table.md")
    with open(out,"w") as fh: fh.write("\n".join(lines)+"\n")
    print("wrote", out, "rows:", len(lines))

if __name__=="__main__": main()
