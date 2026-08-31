#!/usr/bin/env python3
"""Write the 18-seq rendering main table as an .xlsx (reusing the baselines xlsx format
style: one row per seq-method with Method | Dataset | Sequence | ATE | PSNR | SSIM |
LPIPS | Depth-L1 | optional FPS/GPU). Read-only over sources; writes
results/evidence/18seq_rendering_main_table.xlsx.
"""
import openpyxl, statistics
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from build_18seq_main_table import (discover, read_ate, read_summary, SEQORDER, PTYPE,
                                    comp_ate, comp_psnr, comp_ssim, comp_lpips, comp_depth,
                                    read_flowmask)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Main"
ws.append(["Method","Dataset","Sequence","ATE (cm)","PSNR","SSIM","LPIPS","Depth-L1 (cm)"])
# bold header
for cell in ws[1]:
    cell.font = Font(bold=True)

def fmt_our(ate_t, psnr_t, ssim_t, lpips_t, dep_t):
    return [ate_t, psnr_t, ssim_t, lpips_t, dep_t]

disc = discover()
flowmask = read_flowmask()
# For a self-contained xlsx, only include seqs where we have rendering or competitor data.
for seq in SEQORDER:
    d = disc.get(seq, {})
    ptype = PTYPE[seq]
    # our mask-free row
    if "maskfree" in d and d["maskfree"]:
        tsds = d["maskfree"]
        ates = [read_ate(t) for t in tsds.values()]
        ates = [a[0] for a in ates if a]
        # per-metric aggregation
        def agg(k, f1, nd):
            vals = [read_summary(t) for t in tsds.values()]
            vals = [v[k] for v in vals if v and v.get(k) is not None]
            if not vals: return "—"
            m = statistics.mean(vals)
            if len(vals)>1:
                return f"{f1(m-statistics.stdev(vals)/2):.{nd}f}±{statistics.stdev(vals):.{nd}f}"  # placeholder; we want m±sd
            return f"{m:.{nd}f}"
        # simpler: reuse main-table string
        ate = "—"
        if ates:
            m = statistics.mean(ates)
            ate = f"{m:.2f}" + (f"±{statistics.stdev(ates):.2f}" if len(ates)>1 else "")
        ps = [read_summary(t).get("psnr") for t in tsds.values() if read_summary(t)]
        ps=[v for v in ps if v is not None]
        ss=[read_summary(t).get("ssim") for t in tsds.values() if read_summary(t)]; ss=[v for v in ss if v is not None]
        ll=[read_summary(t).get("lpips") for t in tsds.values() if read_summary(t)]; ll=[v for v in ll if v is not None]
        dp=[read_summary(t).get("depth_l1_cm") for t in tsds.values() if read_summary(t)]; dp=[v for v in dp if v is not None]
        def mf(vals, nd):
            if not vals: return "—"
            m=statistics.mean(vals)
            return f"{m:.{nd}f}" + (f"±{statistics.stdev(vals):.{nd}f}" if len(vals)>1 else "")
        ws.append(["Ours-mask-free","BONN/TUM",seq, ate, mf(ps,2), mf(ss,3), mf(ll,3), mf(dp,2)])
    if "maskon" in d and d["maskon"]:
        tsds=d["maskon"]
        def colstate(k):
            vals=[read_summary(t).get(k) for t in tsds.values() if read_summary(t)]
            return [v for v in vals if v is not None]
        ates=[read_ate(t) for t in tsds.values()]; ates=[a[0] for a in ates if a]
        atea="—"
        if ates:
            m=statistics.mean(ates)
            atea=f"{m:.2f}" + (f"±{statistics.stdev(ates):.2f}" if len(ates)>1 else "")
        def mf(vals,nd):
            if not vals: return "—"
            m=statistics.mean(vals)
            return f"{m:.{nd}f}" + (f"±{statistics.stdev(vals):.{nd}f}" if len(vals)>1 else "")
        ws.append(["Ours-combined(mask-ON)","BONN/TUM",seq, atea, mf(colstate("psnr"),2), mf(colstate("ssim"),3), mf(colstate("lpips"),3), mf(colstate("depth_l1_cm"),2)])
    # WP-B naive flow-threshold baseline — held-out 4 seqs only, ATE only (never rendered).
    # Other 14 seqs are absent by campaign scope, documented in the md coverage note.
    if seq in flowmask:
        ws.append(["Baseline-flow-mask(p90)","BONN",seq, flowmask[seq], "—", "—", "—", "—"])
    # competitors
    for m in ["MonoGS","RGD-SLAM"]:
        av=comp_ate.get(m,{}).get(seq)
        pv=comp_psnr.get(m,{}).get(seq); sv=comp_ssim.get(m,{}).get(seq)
        lv=comp_lpips.get(m,{}).get(seq); dv=comp_depth.get(m,{}).get(seq)
        def f(v,nd=2): return "—" if v is None else f"{v:.{nd}f}"
        ate_c = "—" if av is None else f"{av:.2f}"
        ws.append([m,"TUM/BONN",seq, ate_c, f(pv), f(sv,3), f(lv,3), f(dv)])

out="results/evidence/18seq_rendering_main_table.xlsx"
wb.save(out)
print("wrote", out)

if __name__=="__main__":
    pass
